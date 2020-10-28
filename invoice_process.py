import math
import datetime

from openpyxl import load_workbook

from utilities import *


def validate_file(invoice_path, test_mode=False):
    log_file_path = datetime.datetime.now().strftime("logs/%Y%m%d-%H%M%S.txt")
    log_file = open(log_file_path, "w")
    result = True
    print("File Path:", invoice_path, '\n', file=log_file)

    facility = get_facility(invoice_path)
    source = get_source(invoice_path)
    facility_pharmacy_map = get_pharmacy(facility)
    pharmacy = facility_pharmacy_map.pharmacy
    invoice_reader_settings = get_reader_settings(pharmacy, source)

    if not invoice_reader_settings:
        print("Reader setting is not available", file=log_file)
        log_file.close()
        return False, log_file_path, None

    # download invoice
    if not test_mode:
        file_name = 'invoice.xlsx'
        invoice_path = get_s3_client().download_file(get_s3_bucket(), invoice_path, file_name)
    else:
        file_name = 'sample_invoices/' + invoice_path.split('/')[-1]

    invoice_dt = datetime.datetime.now().date()

    # parse invoice
    wb = load_workbook(file_name)

    try:
        sheet_name = invoice_reader_settings.sheet_name or wb.sheetnames[0]
        ws = wb[sheet_name]
    except Exception as e:
        print(f"Required sheet ('{sheetName}') not found.", file=log_file)
        log_file.close()
        return False, log_file_path, None

    # get meta info from [pharmacy_invoice_reader_settings]
    start_index = invoice_reader_settings.header_row_index + invoice_reader_settings.skip_rows_after_header + 1
    nrows = get_valid_rows_count(ws) - invoice_reader_settings.skip_ending_rows
    # import pdb; pdb.set_trace()
    header = [get_clean_header_column(ii.value) for ii in ws[invoice_reader_settings.header_row_index+1] if ii.value]

    # for field in invoice_reader_settings.raw_invoice_fields:
    #     if field.sheet_column_name not in header and not field.is_optional:
    #         result = False
    #         print(f"Column '{field.sheet_column_name}' not found", file=log_file)

    data = []
    # validate each row using field validator
    for row_idx in range(start_index, nrows):
        is_valid, cleaned_data = validate_row(invoice_reader_settings.raw_invoice_fields, header, ws[row_idx+1], row_idx+1, log_file)
        if is_valid:
            data.append(cleaned_data)
        else:
            result = False

    invoice_info = (facility_pharmacy_map, invoice_dt, source, data)
    log_file.close()

    return result, log_file_path, invoice_info


def process_invoice(invoice_info, log_path, test_mode=False):
    (facility_pharmacy_map, invoice_dt, source, invoice_data) = invoice_info
    log_file = open(log_path, 'a')

    # create a log
    source_id = source.id if source else 0
    source_name = source.source_nm.lower() if source else 'general'
    invoice_batch_log_id = start_batch_logging(facility_pharmacy_map, invoice_dt, source_id)
    pharmacy_name = facility_pharmacy_map.pharmacy.pharmacy_nm.lower().replace(' ', '_')
    pharmacy_id = facility_pharmacy_map.pharmacy.id
    facility_id = facility_pharmacy_map.facility.id
    process_invoice_func = globals().get(f'_process_row_{pharmacy_name}_{source_name}')

    result = process_invoice_func(
        invoice_data,
        invoice_batch_log_id,
        pharmacy_id,
        facility_id,
        invoice_dt,
        source_id,
        log_file,
        test_mode
    )
    
    res = stop_batch_logging(invoice_batch_log_id)
    log_file.close()

    return result


def _process_row_speciality_rx_email(invoice_data, invoice_batch_log_id, pharmacy_id, facility_id, invoice_dt, source, log_file, test_mode=False):
    num_inserted = 0
    for row in invoice_data:
        try:
            first_nm = get_first_name(row['patient'])
            last_nm = get_last_name(row['patient'])
            payer_group_id = get_payer_group(pharmacy_id, row['invgrp'], source)
            ssn = row['ssn_no'][:3]+row['ssn_no'][4:6]+row['ssn_no'][7:11] if row['ssn_no'] and row['ssn_no'][0] != '_' else 0

            record = {
                'invoice_batch_id': invoice_batch_log_id,
                'pharmacy_id': pharmacy_id,
                'facility_id': facility_id,
                'payer_group_id': payer_group_id,
                'invoice_dt': invoice_dt,
                'first_nm': first_nm,
                'last_nm': last_nm,
                'ssn': ssn,
                'dob': None,
                'gender': None,
                'dispense_dt': row['dispdt'],
                'product_category': row['rx_otc'],
                'drug_nm': row['drug'],
                'doctor': None,
                'rx_nbr': row['rx_no'],
                'ndc': row['ndc'],
                'reject_cd': None,
                'quantity': row['qty'],
                'days_supplied': row['ds'],
                'charge_amt': row['billamt'],
                'copay_amt': None,
                'copay_flg': 'Y' if row['copay'].upper() == 'COPAY' else None,
                'census_match_cd': None,
                'status_cd': None,
                'charge_confirmed_flg': None,
                'duplicate_flg': test_mode,
                'note': row['comment'],
                'request_credit_flg': None,
                'credit_request_dt': None,
                'credit_request_cd': None,
                'days_overbilled': None
            }

            pharmacy_invoice = PharmacyInvoice(**record)
            session.add(pharmacy_invoice)
            session.commit()
            num_inserted += 1
        except Exception as e:
            print(str(e), file=log_file)
            session.rollback()

    return len(invoice_data) == num_inserted


def _process_row_speciality_rx_portal(invoice_data, invoice_batch_log_id, pharmacy_id, facility_id, invoice_dt, source, log_file, test_mode=False):
    num_inserted = 0
    for row in invoice_data:
        try:
            first_nm = get_first_name(row['resident'])
            last_nm = get_last_name(row['resident'])
            payer_group_id = get_payer_group(pharmacy_id, row['group'], source)
            ssn = row['ssn_no'][:3]+row['ssn_no'][4:6]+row['ssn_no'][7:11] if row['ssn_no'] and row['ssn_no'][0] != '_' else 0

            record = {
                'invoice_batch_id': invoice_batch_log_id,
                'pharmacy_id': pharmacy_id,
                'facility_id': facility_id,
                'payer_group_id': payer_group_id,
                'invoice_dt': invoice_dt,
                'first_nm': first_nm,
                'last_nm': last_nm,
                'ssn': None,
                'dob': None,
                'gender': None,
                'dispense_dt': row['dispensed'],
                'product_category': row['rx_type'],
                'drug_nm': row['drug_nm'],
                'doctor': None,
                'rx_nbr': row['rx_no'],
                'ndc': None,
                'reject_cd': None,
                'quantity': row['quantity'],
                'days_supplied': row['days_supply'],
                'charge_amt': row['amount'],
                'copay_amt': None,
                'copay_flg': 'Y' if row['is_a_copay'].upper() == 'COPAY' else None,
                'census_match_cd': None,
                'status_cd': None,
                'charge_confirmed_flg': None,
                'duplicate_flg': test_mode,
                'note': row['billing_comment'],
                'request_credit_flg': None,
                'credit_request_dt': None,
                'credit_request_cd': None,
                'days_overbilled': None
            }

            pharmacy_invoice = PharmacyInvoice(**record)
            session.add(pharmacy_invoice)
            session.commit()
            num_inserted += 1
        except Exception as e:
            session.rollback()
            print(str(e), file=log_file)

    return len(invoice_data) == num_inserted


def _process_row_pharmscripts_portal(invoice_data, invoice_batch_log_id, pharmacy_id, facility_id, invoice_dt, source, log_file, test_mode=False):
    num_inserted = 0
    for row in invoice_data:
        try:
            first_nm = get_first_name(row['patient_nm'])
            last_nm = get_last_name(row['patient_nm'])
            payer_group_id = get_payer_group(pharmacy_id, row['inv_grp'], source)

            record = {
                'invoice_batch_id': invoice_batch_log_id,
                'pharmacy_id': pharmacy_id,
                'facility_id': facility_id,
                'payer_group_id': payer_group_id,
                'invoice_dt': invoice_dt,
                'first_nm': first_nm,
                'last_nm': last_nm,
                'ssn': row['ssn'],
                'dob': None,
                'gender': 'M' if row['b_or_g'] == 'B' else 'F' if row['b_or_g'] == 'G' else None,
                'dispense_dt': row['disp_dt'],
                'product_category': row['rx_type'],
                'drug_nm': row['drug'],
                'doctor': row['physician'],
                'rx_nbr': row['rx_no'],
                'ndc': row['ndc'],
                'reject_cd': None,
                'quantity': row['qty'],
                'days_supplied': row['ds'],
                'charge_amt': row['bill'],
                'copay_amt': None,
                'copay_flg': 'Y' if row['copay'].upper() == 'Y' else None,
                'census_match_cd': None,
                'status_cd': None,
                'charge_confirmed_flg': None,
                'duplicate_flg': test_mode,
                'note': row['billing_comment'],
                'request_credit_flg': None,
                'credit_request_dt': None,
                'credit_request_cd': None
            }

            pharmacy_invoice = PharmacyInvoice(**record)
            session.add(pharmacy_invoice)
            session.commit()
            num_inserted += 1
        except Exception as e:
            session.rollback()
            print(str(e), file=log_file)

    return len(invoice_data) == num_inserted


def _process_row_pharmscripts_email(invoice_data, invoice_batch_log_id, pharmacy_id, facility_id, invoice_dt, source, log_file, test_mode=False):
    num_inserted = 0
    for row in invoice_data:
        try:
            first_nm = get_first_name(row['patient'])
            last_nm = get_last_name(row['patient'])
            payer_group_id = get_payer_group(pharmacy_id, row['inv_grp'], source)

            record = {
                'invoice_batch_id': invoice_batch_log_id,
                'pharmacy_id': pharmacy_id,
                'facility_id': facility_id,
                'payer_group_id': payer_group_id,
                'invoice_dt': invoice_dt,
                'first_nm': first_nm,
                'last_nm': last_nm,
                'ssn': row['ssn'],
                'dob': None,
                'gender': 'M' if row['b_g'] == 'B' else 'F' if row['b_g'] == 'G' else None,
                'dispense_dt': row['disp_dt'],
                'product_category': row['otc_rx'],
                'drug_nm': row['drug'],
                'doctor': row['physician'],
                'rx_nbr': row['rx_no'],
                'ndc': row['ndc'],
                'reject_cd': None,
                'quantity': row['tot_qty_disp'],
                'days_supplied': row['ds'],
                'charge_amt': row['tot_bill_amt'],
                'copay_amt': None,
                'copay_flg': 'Y' if row['is_a_copay'].upper() == 'Y' else None,
                'census_match_cd': None,
                'status_cd': None,
                'charge_confirmed_flg': None,
                'duplicate_flg': test_mode,
                'note': None,
                'request_credit_flg': None,
                'credit_request_dt': None,
                'credit_request_cd': None,
                'days_overbilled': None
            }

            pharmacy_invoice = PharmacyInvoice(**record)
            session.add(pharmacy_invoice)
            session.commit()
            num_inserted += 1
        except Exception as e:
            session.rollback()
            print(str(e), file=log_file)

    return len(invoice_data) == num_inserted


def _process_row_geriscript_general(invoice_data, invoice_batch_log_id, pharmacy_id, facility_id, invoice_dt, source, log_file, test_mode=False):
    num_inserted = 0
    for row in invoice_data:
        try:
            first_nm = get_first_name(row['full_nm'])
            last_nm = get_last_name(row['full_nm'])
            payer_group_id = get_payer_group(pharmacy_id, row['invoice_grp'], None)
            ssn = row['ssn'][:3]+row['ssn'][4:6]+row['ssn'][7:11] if row['ssn'] and row['ssn'][0] != '_' else ''

            record = {
                'invoice_batch_id': invoice_batch_log_id,
                'pharmacy_id': pharmacy_id,
                'facility_id': facility_id,
                'payer_group_id': payer_group_id,
                'invoice_dt': invoice_dt,
                'first_nm': first_nm,
                'last_nm': last_nm,
                'ssn': ssn,
                'dob': row['birth_date'],
                'gender': row['sex'],
                'dispense_dt': row['dispense_dt'],
                'product_category': row['rx_otc'],
                'drug_nm': row['drug_label_nm'],
                'doctor': row['doctor'],
                'rx_nbr': row['rx_no'],
                'ndc': row['ndc'],
                'reject_cd': None,
                'quantity': row['qty'],
                'days_supplied': row['days_supply'],
                'charge_amt': row['bill_amt'],
                'copay_amt': row['copay_amt'],
                'copay_flg': 'Y' if row['copay_amt'] > 0 else None,
                'census_match_cd': None,
                'status_cd': None,
                'charge_confirmed_flg': None,
                'duplicate_flg': test_mode,
                'note': row['billing_comment'],
                'request_credit_flg': None,
                'credit_request_dt': None,
                'credit_request_cd': None,
                'days_overbilled': None
            }

            pharmacy_invoice = PharmacyInvoice(**record)
            session.add(pharmacy_invoice)
            session.commit()
            num_inserted += 1
        except Exception as e:
            session.rollback()
            print(str(e), file=log_file)

    return len(invoice_data) == num_inserted


def _process_row_medwiz_general(invoice_data, invoice_batch_log_id, pharmacy_id, facility_id, invoice_dt, source, log_file, test_mode=False):
    num_inserted = 0
    for row in invoice_data:
        try:
            first_nm = get_first_name(row['name'])
            last_nm = get_last_name(row['name'])
            payer_group_id = get_payer_group(pharmacy_id, row['invoice_group'], None)

            record = {
                'invoice_batch_id': invoice_batch_log_id,
                'pharmacy_id': pharmacy_id,
                'facility_id': facility_id,
                'payer_group_id': payer_group_id,
                'invoice_dt': invoice_dt,
                'first_nm': first_nm,
                'last_nm': last_nm,
                'ssn': None,
                'dob': None,
                'gender': None,
                'dispense_dt': row['dispense_date'],
                'product_category': row['distribution_code'],
                'drug_nm': row['description'],
                'doctor': None,
                'rx_nbr': row['rx_no'],
                'ndc': row['ndc'],
                'reject_cd': None,
                'quantity': row['qty'],
                'days_supplied': row['days_supply'],
                'charge_amt': row['amount'],
                'copay_amt': None,
                'copay_flg': 'Y' if row['copay'].upper() == 'COPAY' else None,
                'census_match_cd': None,
                'status_cd': None,
                'charge_confirmed_flg': None,
                'duplicate_flg': test_mode,
                'note': row['billing_comment'],
                'request_credit_flg': None,
                'credit_request_dt': None,
                'credit_request_cd': None,
                'days_overbilled': None
            }

            pharmacy_invoice = PharmacyInvoice(**record)
            session.add(pharmacy_invoice)
            session.commit()
            num_inserted += 1
        except Exception as e:
            session.rollback()
            print(str(e), file=log_file)

    return len(invoice_data) == num_inserted


def _process_row_omnicare_general(invoice_data, invoice_batch_log_id, pharmacy_id, facility_id, invoice_dt, source, log_file, test_mode=False):
    num_inserted = 0
    for row in invoice_data:
        try:
            first_nm = row['patient_first_nm']
            last_nm = row['patient_last_nm']
            payer_group_id = get_payer_group(pharmacy_id, row['pay_type_description'], None)
            ssn = row['patient_ssn'][:3]+row['patient_ssn'][4:6]+row['patient_ssn'][7:11] if row['patient_ssn'] and row['patient_ssn'][0] != '_' else ''

            record = {
                'invoice_batch_id': invoice_batch_log_id,
                'pharmacy_id': pharmacy_id,
                'facility_id': facility_id,
                'payer_group_id': payer_group_id,
                'invoice_dt': invoice_dt,
                'first_nm': first_nm,
                'last_nm': last_nm,
                'ssn': ssn,
                'dob': None,
                'gender': None,
                'dispense_dt': row['transaction_dt'],
                'product_category': row['inventory_category'],
                'drug_nm': row['description'],
                'doctor': row['physician'],
                'rx_nbr': row['rx'],
                'ndc': row['ndc'],
                'reject_cd': row['reject_codes'],
                'quantity': row['qty'],
                'days_supplied': row['days_supply'],
                'charge_amt': row['amount'],
                'copay_amt': row['amount'] if row['copay'] == 'copay' else None,
                'copay_flg': 'Y' if row['copay'] == 'copay' else None,
                'census_match_cd': None,
                'status_cd': None,
                'charge_confirmed_flg': None,
                'duplicate_flg': test_mode,
                'note': row['statement_note'],
                'request_credit_flg': None,
                'credit_request_dt': None,
                'credit_request_cd': None,
                'days_overbilled': None
            }

            pharmacy_invoice = PharmacyInvoice(**record)
            session.add(pharmacy_invoice)
            session.commit()
            num_inserted += 1
        except Exception as e:
            print(str(e), file=log_file)
            session.rollback()

    return len(invoice_data) == num_inserted


def _process_row_pharmerica_email(invoice_data, invoice_batch_log_id, pharmacy_id, facility_id, invoice_dt, source, log_file, test_mode=False):
    num_inserted = 0
    for row in invoice_data:
        try:
            first_nm = get_first_name(row['resident_nm'])
            last_nm = get_last_name(row['resident_nm'])
            payer_group_id = get_payer_group(pharmacy_id, row['fin_plan'], source)
            ssn = row['res_ssn'][:3]+row['res_ssn'][4:6]+row['res_ssn'][7:11] if row['res_ssn'] and row['res_ssn'][0] != '_' else ''

            record = {
                'invoice_batch_id': invoice_batch_log_id,
                'pharmacy_id': pharmacy_id,
                'facility_id': facility_id,
                'payer_group_id': payer_group_id,
                'invoice_dt': invoice_dt,
                'first_nm': first_nm,
                'last_nm': last_nm,
                'ssn': ssn,
                'dob': None,
                'gender': None,
                'dispense_dt': row['service_dt'],
                'product_category': ((row['product_category'] or '') + (row['sales_type'] or '')) or None,
                'drug_nm': row['trans_desc'],
                'doctor': row['doctor_nm'],
                'rx_nbr': row['rx_nbr'],
                'ndc': row['ndc_nbr'],
                'reject_cd': None,
                'quantity': row['quantity'],
                'days_supplied': row['days_supply'],
                'charge_amt': row['amount_due'],
                'copay_amt': None,
                'copay_flg': None,
                'census_match_cd': None,
                'status_cd': None,
                'charge_confirmed_flg': None,
                'duplicate_flg': test_mode,
                'note': row['task_manager_notes'],
                'request_credit_flg': None,
                'credit_request_dt': None,
                'credit_request_cd': None,
                'days_overbilled': None
            }

            pharmacy_invoice = PharmacyInvoice(**record)
            session.add(pharmacy_invoice)
            session.commit()
            num_inserted += 1
        except Exception as e:
            session.rollback()
            print(str(e), file=log_file)

    return len(invoice_data) == num_inserted


def _process_row_pharmerica_portal(invoice_data, invoice_batch_log_id, pharmacy_id, facility_id, invoice_dt, source, log_file, test_mode=False):
    num_inserted = 0
    for row in invoice_data:
        try:
            first_nm = get_first_name(row['resident_nm'])
            last_nm = get_last_name(row['resident_nm'])
            payer_group_id = get_payer_group(pharmacy_id, row['fin_plan'], source)
            ssn = row['res_ssn'][:3]+row['res_ssn'][4:6]+row['res_ssn'][7:11] if row['res_ssn'] and row['res_ssn'][0] != '_' else ''

            record = {
                'invoice_batch_id': invoice_batch_log_id,
                'pharmacy_id': pharmacy_id,
                'facility_id': facility_id,
                'payer_group_id': payer_group_id,
                'invoice_dt': invoice_dt,
                'first_nm': first_nm,
                'last_nm': last_nm,
                'ssn': ssn,
                'dob': None,
                'gender': None,
                'dispense_dt': row['service_dt'],
                'product_category': row['product_category'],
                'drug_nm': row['trans_desc'],
                'doctor': row['doctor_nm'],
                'rx_nbr': math.floor(row['rx_nbr']),
                'ndc': row['ndc_nbr'],
                'reject_cd': None,
                'quantity': row['quantity'],
                'days_supplied': row['days_supply'],
                'charge_amt': row['trans_amount'],
                'copay_amt': None,
                'copay_flg': None,
                'census_match_cd': None,
                'status_cd': None,
                'charge_confirmed_flg': None,
                'duplicate_flg': test_mode,
                'note': row['task_manager_notes'],
                'request_credit_flg': None,
                'credit_request_dt': None,
                'credit_request_cd': None,
                'days_overbilled': None
            }

            pharmacy_invoice = PharmacyInvoice(**record)
            session.add(pharmacy_invoice)
            session.commit()
            num_inserted += 1
        except Exception as e:
            session.rollback()
            print(str(e), file=log_file)

    return len(invoice_data) == num_inserted
